from zipfile import ZipFile
import csv
import os
import shutil
from PyPDF2 import PdfReader
from openpyxl import load_workbook


directory = os.path.join("C:/Users/bohda/PycharmProjects/Working_with_files", "resources")
if not os.path.exists(directory):
    os.mkdir(directory)

with ZipFile('resources/sample.zip', 'w') as myzip:
    myzip.write('docs-pytest-org-en-latest.pdf')
    myzip.write('file_example_XLSX_50.xlsx')
    myzip.write('username.csv')


'''
or

directory = "resources"
parent_dir = "C:/Users/bohda/PycharmProjects/Working_with_files"
path = os.path.join(parent_dir, directory)
os.makedirs(path)

with ZipFile('resources/sample.zip', 'w') as myzip:
    myzip.write('docs-pytest-org-en-latest.pdf')
    myzip.write('file_example_XLSX_50.xlsx')
    myzip.write('username.csv')
    
or
    
os.makedirs('resources')

'''

with ZipFile('resources/sample.zip', 'r') as myzip:
    myzip.extractall('resources')
    myzip.close()


def test_pdf_number_of_pages():
    reader = PdfReader("resources/docs-pytest-org-en-latest.pdf")
    number_of_pages = len(reader.pages)
    print(number_of_pages)
    assert number_of_pages == 412


def test_pdf_search_text():
    reader = PdfReader("resources/docs-pytest-org-en-latest.pdf")
    page = reader.pages[7]
    text = page.extract_text()
    print(text)
    assert "Run multiple tests" in text


def test_csv_size():
    read_csv = open('resources/username.csv')
    reader = csv.reader(read_csv, delimiter=';')
    for row in reader:
        print(row)
    assert os.path.getsize("resources/username.csv") == 181


def test_check_the_first_line():
    read_csv = open('resources/username.csv')
    content = read_csv.readlines()
    rows = content[1]
    print(rows)
    assert 'booker12;9012;Rachel;Booker' in rows


def test_file_xlsx():
    workbook = load_workbook('resources/file_example_XLSX_50.xlsx')
    sheet = workbook.active
    result_value = sheet.cell(row=3, column=5).value
    print(result_value)
    assert 'Great Britain' in result_value


def test_remove_folder():
    shutil.rmtree('resources')